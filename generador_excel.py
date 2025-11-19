import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import statistics

def leer_costo_optimo(archivo_datos, carpeta_datos='Datos'):
    """Lee el costo óptimo de un archivo .sln"""
    try:
        ruta = Path(carpeta_datos) / archivo_datos
        with open(ruta, 'r') as f:
            primera_linea = f.readline().strip()
            partes = primera_linea.split()
            if len(partes) >= 2:
                return int(partes[1])
    except Exception as e:
        print(f"Error leyendo {archivo_datos}: {e}")
    return None

def extraer_info_log(archivo_log):
    """Extrae información relevante del archivo de log"""
    try:
        with open(archivo_log, 'r', encoding='utf-8') as f:
            contenido = f.read()
        
        # Buscar la sección "SOLUCIÓN FINAL" y extraer costo y tiempo de ahí
        # Esto asegura que tomamos los valores finales correctos
        match_solucion = re.search(r'SOLUCIÓN FINAL.*?Costo:\s*(\d+)\s*\|\s*Gen:\s*\d+\s*\|\s*Tiempo:\s*([\d.]+)s', contenido, re.DOTALL)
        
        if match_solucion:
            costo = int(match_solucion.group(1))
            tiempo = float(match_solucion.group(2))
        else:
            # Fallback: buscar cualquier ocurrencia (por compatibilidad)
            match_costo = re.search(r'Costo:\s*(\d+)', contenido)
            costo = int(match_costo.group(1)) if match_costo else None
            
            match_tiempo = re.search(r'Tiempo:\s*([\d.]+)s', contenido)
            tiempo = float(match_tiempo.group(1)) if match_tiempo else None
        
        # Extraer semilla (si existe en el nombre del archivo)
        match_semilla = re.search(r'_(\d{8,})(?:_|\.txt)', str(archivo_log))
        semilla = match_semilla.group(1) if match_semilla else None
        
        return {
            'costo': costo,
            'tiempo': tiempo,
            'semilla': semilla
        }
    except Exception as e:
        print(f"Error procesando {archivo_log}: {e}")
        return None

def agrupar_archivos_logs(carpeta_logs='Logs'):
    """Agrupa los archivos de log por tipo de algoritmo y archivo de datos"""
    grupos = {}
    
    for archivo in Path(carpeta_logs).glob('*.txt'):
        nombre = archivo.stem
        
        # Parsear nombre: tipo_variante_archivoDatos_semilla_infoExtra
        partes = nombre.split('_')
        if len(partes) >= 3:
            tipo = partes[0]  # evolutivo
            variante = partes[1]  # generacional o estacionario
            archivo_datos = partes[2]  # ford01, etc
            
            # Encontrar la semilla (número largo) y separar info_extra
            semilla = None
            indice_semilla = -1
            info_extra_partes = []
            
            for i in range(3, len(partes)):
                # Si es un número de 8+ dígitos, es la semilla
                if partes[i].isdigit() and len(partes[i]) >= 8:
                    semilla = partes[i]
                    indice_semilla = i
                    # Todo lo que viene después de la semilla es info_extra
                    info_extra_partes = partes[i+1:]
                    break
            
            # Si no se encontró semilla, toda la info después del archivo es info_extra
            if semilla is None:
                info_extra_partes = partes[3:]
            
            info_extra = '_'.join(info_extra_partes) if info_extra_partes else ''
            
            # Crear clave de grupo (sin semilla)
            clave = f"{tipo}_{variante}_{info_extra}" if info_extra else f"{tipo}_{variante}"
            
            if clave not in grupos:
                grupos[clave] = {}
            
            if archivo_datos not in grupos[clave]:
                grupos[clave][archivo_datos] = []
            
            grupos[clave][archivo_datos].append(str(archivo))
    
    return grupos

def crear_hoja_resumen(ws, nombre_configuracion, grupo_datos, carpeta_datos='Datos'):
    """Crea la tabla resumen en una hoja de Excel con fórmulas"""
    
    # Obtener archivos de datos únicos y ordenarlos
    archivos_datos = sorted(grupo_datos.keys())
    
    # Calcular el número total de columnas necesarias
    num_columnas_totales = len(archivos_datos) * 2 + 2  # 2 por archivo (Sol, Time) + 1 etiqueta + 1 semilla
    
    # Fila 1: Título de configuración
    # Asegurarse de combinar suficientes celdas
    ultima_col_titulo = max(8, num_columnas_totales)
    ws.merge_cells(f'A1:{get_column_letter(ultima_col_titulo)}1')
    cell_config = ws['A1']
    cell_config.value = nombre_configuracion.upper().replace('_', ' ')
    cell_config.font = Font(bold=True, size=14, color="FFFFFF")
    cell_config.alignment = Alignment(horizontal='center', vertical='center')
    cell_config.fill = PatternFill(start_color="3F9E5E", end_color="3F9E5E", fill_type="solid")
    ws.row_dimensions[1].height = 25
    
    fila_actual = 2
    
    # Fila 2: Nombres de archivos (fusionadas)
    col = 1
    ws.cell(row=fila_actual, column=col).value = ''
    col += 1
    
    for arch in archivos_datos:
        nombre_upper = arch.upper().replace('.SLN', '')
        ws.merge_cells(start_row=fila_actual, start_column=col, 
                      end_row=fila_actual, end_column=col+1)
        cell = ws.cell(row=fila_actual, column=col)
        cell.value = nombre_upper
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="70C78D", end_color="70C78D", fill_type="solid")
        col += 2
    
    ws.cell(row=fila_actual, column=col).value = ''
    fila_actual += 1
    
    # Fila 3: Tamaño
    col = 1
    ws.cell(row=fila_actual, column=col).value = 'GREEDY AL'
    ws.cell(row=fila_actual, column=col).font = Font(bold=True)
    col += 1
    
    for arch_datos in archivos_datos:
        arch_sln = f"{arch_datos}.sln" if not arch_datos.endswith('.sln') else arch_datos
        
        # Obtener tamaño
        try:
            ruta = Path(carpeta_datos) / arch_sln
            with open(ruta, 'r') as f:
                primera_linea = f.readline().strip()
                tamaño = int(primera_linea.split()[0])
        except:
            tamaño = '?'
        
        ws.cell(row=fila_actual, column=col).value = 'Tamaño'
        ws.cell(row=fila_actual, column=col+1).value = tamaño
        col += 2
    
    ws.cell(row=fila_actual, column=col).value = ''
    fila_actual += 1
    
    # Fila 4: Mínimo global
    col = 1
    ws.cell(row=fila_actual, column=col).value = ''
    col += 1
    
    for arch_datos in archivos_datos:
        arch_sln = f"{arch_datos}.sln" if not arch_datos.endswith('.sln') else arch_datos
        costo_optimo = leer_costo_optimo(arch_sln, carpeta_datos)
        
        ws.cell(row=fila_actual, column=col).value = 'Minimo global'
        ws.cell(row=fila_actual, column=col+1).value = costo_optimo if costo_optimo else '?'
        col += 2
    
    ws.cell(row=fila_actual, column=col).value = ''
    fila_actual += 1
    
    # Fila vacía
    fila_actual += 1
    
    # Fila de encabezados
    col = 1
    ws.cell(row=fila_actual, column=col).value = ''
    col += 1
    
    for _ in archivos_datos:
        cell_sol = ws.cell(row=fila_actual, column=col)
        cell_sol.value = 'Sol'
        cell_sol.font = Font(bold=True, italic=True)
        cell_sol.fill = PatternFill(start_color="A8D5BA", end_color="A8D5BA", fill_type="solid")
        
        cell_time = ws.cell(row=fila_actual, column=col+1)
        cell_time.value = 'Time'
        cell_time.font = Font(bold=True, italic=True)
        cell_time.fill = PatternFill(start_color="A8D5BA", end_color="A8D5BA", fill_type="solid")
        col += 2
    
    ws.cell(row=fila_actual, column=col).value = 'Semilla'
    ws.cell(row=fila_actual, column=col).font = Font(bold=True)
    ws.cell(row=fila_actual, column=col).fill = PatternFill(start_color="A8D5BA", end_color="A8D5BA", fill_type="solid")
    
    fila_encabezado = fila_actual
    fila_actual += 1
    
    # Recopilar datos y ordenar por semilla
    datos_por_archivo = {}
    for arch_datos in archivos_datos:
        datos_por_archivo[arch_datos] = []
        
        for log_file in grupo_datos[arch_datos]:
            info = extraer_info_log(log_file)
            if info:
                datos_por_archivo[arch_datos].append(info)
        
        # Ordenar por semilla para mantener consistencia
        datos_por_archivo[arch_datos].sort(key=lambda x: x['semilla'] or '')
    
    max_ejecuciones = max(len(datos) for datos in datos_por_archivo.values())
    
    # Filas de ejecuciones con datos reales
    primera_fila_datos = fila_actual
    for i in range(max_ejecuciones):
        col = 1
        ws.cell(row=fila_actual, column=col).value = f'Ejecución {i+1}'
        ws.cell(row=fila_actual, column=col).font = Font(bold=True)
        col += 1
        
        semilla_fila = None
        
        for arch_datos in archivos_datos:
            datos = datos_por_archivo[arch_datos]
            if i < len(datos):
                ws.cell(row=fila_actual, column=col).value = datos[i]['costo']
                ws.cell(row=fila_actual, column=col+1).value = datos[i]['tiempo']
                if semilla_fila is None:
                    semilla_fila = datos[i]['semilla']
            col += 2
        
        ws.cell(row=fila_actual, column=col).value = semilla_fila if semilla_fila else ''
        fila_actual += 1
    
    ultima_fila_datos = fila_actual - 1
    
    # Fila de desviación típica - calcular directamente en Python
    col = 1
    ws.cell(row=fila_actual, column=col).value = 'Desv. típica'
    ws.cell(row=fila_actual, column=col).font = Font(bold=True, italic=True)
    ws.cell(row=fila_actual, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    col += 1
    
    for idx_archivo, arch_datos in enumerate(archivos_datos):
        # Calcular desviación típica de errores relativos en Python
        datos = datos_por_archivo[arch_datos]
        if datos:
            arch_sln = f"{arch_datos}.sln" if not arch_datos.endswith('.sln') else arch_datos
            costo_optimo = leer_costo_optimo(arch_sln, carpeta_datos)
            if costo_optimo:
                errores_relativos = [(d['costo'] - costo_optimo) / costo_optimo for d in datos if d['costo']]
                if len(errores_relativos) > 1:
                    desv_tipica = statistics.stdev(errores_relativos)
                    ws.cell(row=fila_actual, column=col).value = desv_tipica
                    ws.cell(row=fila_actual, column=col).number_format = '0.00%'
                else:
                    ws.cell(row=fila_actual, column=col).value = ''
            else:
                ws.cell(row=fila_actual, column=col).value = ''
        else:
            ws.cell(row=fila_actual, column=col).value = ''
        
        ws.cell(row=fila_actual, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Columna de Time: vacía
        ws.cell(row=fila_actual, column=col+1).value = ''
        
        col += 2
    
    ws.cell(row=fila_actual, column=col).value = ''
    fila_desv_tipica = fila_actual
    fila_actual += 1
    
    # Fila de desviación de error - calcular directamente en Python
    col = 1
    ws.cell(row=fila_actual, column=col).value = 'Desviación de error'
    ws.cell(row=fila_actual, column=col).font = Font(bold=True, italic=True)
    ws.cell(row=fila_actual, column=col).fill = PatternFill(start_color="A8D5BA", end_color="A8D5BA", fill_type="solid")
    col += 1
    
    for idx_archivo, arch_datos in enumerate(archivos_datos):
        # Calcular promedio de errores relativos en Python
        datos = datos_por_archivo[arch_datos]
        if datos:
            arch_sln = f"{arch_datos}.sln" if not arch_datos.endswith('.sln') else arch_datos
            costo_optimo = leer_costo_optimo(arch_sln, carpeta_datos)
            if costo_optimo:
                errores_relativos = [(d['costo'] - costo_optimo) / costo_optimo for d in datos if d['costo']]
                if errores_relativos:
                    promedio_error = statistics.mean(errores_relativos)
                    ws.cell(row=fila_actual, column=col).value = promedio_error
                    ws.cell(row=fila_actual, column=col).number_format = '0.00%'
                else:
                    ws.cell(row=fila_actual, column=col).value = ''
            else:
                ws.cell(row=fila_actual, column=col).value = ''
        else:
            ws.cell(row=fila_actual, column=col).value = ''
        
        ws.cell(row=fila_actual, column=col).fill = PatternFill(start_color="A8D5BA", end_color="A8D5BA", fill_type="solid")
        
        # Columna Time: vacía
        ws.cell(row=fila_actual, column=col+1).value = ''
        
        col += 2
    
    ws.cell(row=fila_actual, column=col).value = ''
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 18
    for i in range(2, len(archivos_datos) * 2 + 2):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions[get_column_letter(len(archivos_datos) * 2 + 2)].width = 15
    
    # Aplicar bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=2, max_row=fila_actual, 
                           min_col=1, max_col=len(archivos_datos)*2+2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

def generar_excel(carpeta_logs='Logs', carpeta_datos='Datos', archivo_salida='resumen.xlsx'):
    """Genera el archivo Excel con todas las tablas resumen"""
    
    # Agrupar archivos
    grupos = agrupar_archivos_logs(carpeta_logs)
    
    if not grupos:
        print("No se encontraron archivos de log para procesar.")
        return
    
    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Crear una hoja por cada grupo
    for nombre_grupo, datos_grupo in grupos.items():
        # Crear hoja con nombre apropiado (limitado a 31 caracteres para Excel)
        # Reemplazar caracteres problemáticos
        nombre_hoja = nombre_grupo[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_').replace('[', '_').replace(']', '_').replace(':', '_')
        
        # Asegurar que el nombre sea único
        nombre_base = nombre_hoja
        contador = 1
        while nombre_hoja in [sheet.title for sheet in wb.worksheets]:
            nombre_hoja = f"{nombre_base[:28]}_{contador}"
            contador += 1
        
        ws = wb.create_sheet(title=nombre_hoja)
        
        # Crear tabla resumen con fórmulas
        crear_hoja_resumen(ws, nombre_grupo, datos_grupo, carpeta_datos)
    
    # Guardar archivo
    try:
        wb.save(archivo_salida)
        print(f"✓ Excel generado exitosamente: {archivo_salida}")
        print(f"✓ Se crearon {len(grupos)} hojas de configuración")
    except Exception as e:
        print(f"✗ Error al guardar el archivo: {e}")

# Ejecutar
if __name__ == "__main__":
    generar_excel(
        carpeta_logs='Logs',
        carpeta_datos='Datos',
        archivo_salida='resumen_experimentos.xlsx'
    )